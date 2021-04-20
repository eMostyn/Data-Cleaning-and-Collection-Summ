import urllib.request
import urllib.parse
from bs4 import BeautifulSoup
import os
import math
import openpyxl
import xlsxwriter as xls
import numpy as np
import time
import matplotlib.pyplot as plt
import seaborn
import pandas as pd
#keywords = ["targeted threat","Advanced Persistent Threat","phishing","DoS attack","malware","computer virus","spyware","malicious bot","ransomware","encryption"]

            
def problem1():
    keywords = []
    keywordsBook = openpyxl.load_workbook(filename = "keywords.xlsx")
    sheet=keywordsBook.active
    yPos = 2
    newKeyword = sheet['A2'].value
    while(newKeyword != None):
        keywords.append(newKeyword)
        yPos+=1
        newKeyword = sheet['A'+str(yPos)].value
        print(newKeyword)
    print(keywords)
    
    url = "https://www.bbc.co.uk/search"
    articles = []
    for keyword in keywords:
        print(keyword)
        keywordArticles = []
        page = 0
        moreResults = True
        while len(keywordArticles) < 100 and moreResults:
            moreResults = False
            page += 1
            values = {'q':keyword,'page':str(page)}
            #print(values)
            data = urllib.parse.urlencode(values)
            req = urllib.request.Request(url+'?'+data)
            resp = urllib.request.urlopen(req)
            respData = resp.read()
            string = "Search results for "+keyword+"</h2>"
            test = string.encode('utf_8')
            searchData = respData[respData.find(test):]
            soup = BeautifulSoup(searchData,features="html.parser")
            links = soup.find_all('a',attrs={'class':'ssrcss-1b5rnkt-PromoLink e1f5wbog5'})
            #print(links)
            for link in links:
                result = link.get('href')
                if  len(keywordArticles) == 100:
                    moreResults = False
                    break
                elif "news" in result or "sport" in result:
                #elif "programme" not in result and "/sound" not in result and "/search"  not in result and "/iplayer" not in result and "/guide" not in result and "/blogs" not in result /learning_english ,/education:
                    keywordArticles.append(result)
                    moreResults = True
                    #print(result)
        articles.append(keywordArticles)

    for i in range(0,len(articles)):
        print(keywords[i],len(articles[i]))
        
    keywordData = []
    keywordIndex = -1
    for keywordArticles in articles:
        keywordIndex += 1
        articleData = []
        file = open(keywords[keywordIndex]+"/"+"urls.txt", "w",encoding='utf-8')
        i=0
        for article in keywordArticles:
            i+=1
            url =  article
            file.write(str(i)+": "+url+"\n")
            #print(url)
            req = urllib.request.Request(url)
            resp = urllib.request.urlopen(req)
            data = resp.read()
            articleData.append(data)
        keywordData.append(articleData)

    return keywordData,keywords

def problem2(articles,keywords):
    for keyword in keywords:
        if os.path.exists(os.path.join(os.getcwd(), keyword))==False:
            os.mkdir(keyword)
            
    
    keywordNumber=-1
    for keyword in articles:
        keywordNumber+=1
        keywordData = []
        articleNumber = 0
        for article in keyword:
            articleNumber +=1
            fileName = keywords[keywordNumber]+str(articleNumber)+".txt"
            file = open(keywords[keywordNumber]+"/"+fileName, "w",encoding='utf-8')
            file.write(fileName+"\n")
            soup = BeautifulSoup(article,features="html.parser")
            paragraphs = soup.find_all('p')
            #divs = soup.findAll('div',attrs={"class":"ssrcss-3z08n3-RichTextContainer e5tfeyi2"})
            textElements = [item for item in soup.findAll(text=True)if item.parent.name == 'p']
            for element in textElements:
                if articleNumber == 21 and keywordNumber ==5:
                    print(element)
                file.write(element)
            file.close()
    



def problem3(keywords):
    keywordVectors = []
    keywordDocuments = []
   # words = []
    for keyword in keywords:
        keywordVector = {}
        keywordDocument = []
        num = 1
        #fileName = os.path.join(os.getcwd(),keyword)
        fileName =  os.path.join(os.getcwd(),keyword,keyword+str(num)+".txt")
        print(fileName)
        while(os.path.isfile(fileName)):
            article = open(fileName,"r",encoding='utf-8')
            data = article.read()
            keywordDocument.append(data)
            keywordVector.update(getVector(data))
            num += 1
            fileName = os.path.join(os.getcwd(),keyword,keyword+str(num)+".txt")
        keywordVectors.append(keywordVector)
        #words.extend(keywordVector.keys())
        keywordDocuments.append(keywordDocument)
    start = time.time()

    similarities = []
    for i in range(0,len(keywordVectors)-1):
        keywordSim = []
        for j in range(i+1,len(keywordVectors)):
            sim = pearsonSimilarity(keywordVectors[i],keywordVectors[j],keywordDocuments[i]+keywordDocuments[j])
            print("Similarity between: "+keywords[i] + " & "+keywords[j]+ " "+ str(sim))
            keywordSim.append(sim)
        similarities.append(keywordSim)

    #print(similarities)

    workbook = xls.Workbook('distance.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write(0,0,"Keywords")
    xPos = 0
    for i in range(0,len(keywords)):
        worksheet.write(0,i+1,keywords[i])
        worksheet.write(i+1,0,keywords[i])
        worksheet.write(i+1, i+1, 0)
    startingX = 0
    xPos = 0
    yPos = 0
    for i in range(0,len(similarities)):
        yPos += 1
        startingX += 1
        xPos = startingX
        for j in range(0,len(similarities[i])):
            xPos +=  1
            #print(similarities[i][j])
            worksheet.write(yPos, xPos, similarities[i][j])
            worksheet.write(xPos, yPos, similarities[i][j])     
    workbook.close()


def getVector(inputData):
    keywordVector = {}
    words = inputData.split(" ")
    for word in words:
            if word in keywordVector.keys():
                keywordVector[word] += 1
            else:
                keywordVector[word] = 1
    return keywordVector

def getSim(x,y,documents):
    mergedDict = {**x, **y}
    comparisons = []
    idf = inverseDocumentFrequency(mergedDict.keys(),documents)
    for word in mergedDict.keys():
        toAdd = []
        if word in x:
            toAdd.append(x[word]*idf[word])
        else:
            toAdd.append(0)
        if word in y:
            toAdd.append(y[word]*idf[word])
        else:
            toAdd.append(0)
        comparisons.append(toAdd)
    
    top = 0
    xTotal = 0
    yTotal = 0
    for pair in comparisons:
        top += pair[0]*pair[1]
        xTotal +=pair[0]**2
        yTotal += pair[1]**2
    xTotal = math.sqrt(xTotal)
    yTotal = math.sqrt(yTotal)
    similarity = top/(xTotal*yTotal)
    return similarity
        
def pearsonSimilarity(x,y,documents):
    mergedDict = {**x, **y}
    idf = inverseDocumentFrequency(mergedDict.keys(),documents)
    comparisons = []
    for word in mergedDict.keys():
        toAdd = []
        if word in x:
            toAdd.append(x[word]*idf[word])
        else:
            toAdd.append(0)
        if word in y:
            toAdd.append(y[word]*idf[word])
        else:
            toAdd.append(0)
        comparisons.append(toAdd)
    comparisons = np.array(comparisons)
    xMean = np.mean(comparisons[:,0])
    yMean = np.mean(comparisons[:,1])
    x  = comparisons[:,0]-xMean
    y = comparisons[:,1] -yMean
    numer = np.sum(x*y)
    denom = np.sqrt( (np.sum(x**2) ) *  (np.sum(y**2) ))
    return 1-(numer/denom)

def inverseDocumentFrequency(dictionary, documents):
    idf = {}
    for word in dictionary:
        idf[word] = 0
        for document in documents:
            if word in document:
                idf[word] += 1
        idf[word] = math.log(1+(len(documents)/idf[word]))
    return idf

def problem4():
    data = pd.read_excel('distance.xlsx',header = 0,index_col=0)
    #data.set_index('Keywords')
    print(data)
    
    seaborn.set_theme()
    # Load the example flights dataset and convert to long-form
    flights_long = seaborn.load_dataset("flights")
    flights = flights_long.pivot("month", "year", "passengers")
    print(flights)
    

    seaborn.heatmap(data,vmin = 0,vmax = 2);
    # Draw a heatmap with the numeric values in each cell

##    seaborn.heatmap(flights, annot=True, fmt="d", linewidths=.5, ax=ax)
    plt.show()

def test():
    keywords = []
    keywordsBook = openpyxl.load_workbook(filename = "keywords.xlsx")
    sheet=keywordsBook.active
    yPos = 2
    newKeyword = sheet['A2'].value
    while(newKeyword != None):
        keywords.append(newKeyword)
        yPos+=1
        newKeyword = sheet['A'+str(yPos)].value
    print(keywords)
    
    url = "https://edition.cnn.com/search"
    articles = []
    for keyword in keywords:
        print(keyword)
        keywordArticles = []
        page = 0
        moreResults = True
        while len(keywordArticles) < 100 and moreResults:
            moreResults = False
            page += 1
            values = {'q':keyword}
            #print(values)
            data = urllib.parse.urlencode(values)
            req = urllib.request.Request(url+'?'+data)
            print(url+'?'+data)
            resp = urllib.request.urlopen(req)
            respData = resp.read()
            print(respData)
            
            soup = BeautifulSoup(respData,features="html.parser")
            divs = soup.findAll('div',attrs={'class':'cnn-search__result-contents'})
            #links = soup.find_all('a')
            for div in divs:
                print("here")
                link = div.findall('a')
                result = link.get('href')
                print(result)
                return
                if  len(keywordArticles) == 100:
                    moreResults = False
                    break
                elif "news" in result or "sport" in result:
                #elif "programme" not in result and "/sound" not in result and "/search"  not in result and "/iplayer" not in result and "/guide" not in result and "/blogs" not in result /learning_english ,/education:
                    keywordArticles.append(result)
                    moreResults = True
                    #print(result)
        articles.append(keywordArticles)

    for i in range(0,len(articles)):
        print(keywords[i],len(articles[i]))
        
    keywordData = []
    keywordIndex = -1
    for keywordArticles in articles:
        keywordIndex += 1
        articleData = []
        file = open(keywords[keywordIndex]+"/"+"urls.txt", "w",encoding='utf-8')
        i=0
        for article in keywordArticles:
            i+=1
            url =  article
            file.write(str(i)+": "+url+"\n")
            #print(url)
            req = urllib.request.Request(url)
            resp = urllib.request.urlopen(req)
            data = resp.read()
            articleData.append(data)
        keywordData.append(articleData)

    return keywordData,keywords
articleData,keywords = problem1()
print("done")
problem2(articleData, keywords)
print("done2")
problem3(keywords)   
problem4()
#test()
